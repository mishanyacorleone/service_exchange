import io
from openpyxl import load_workbook
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from accounts.models import UserProfile
from orders.models import Order, Bid
from accounts.admin import UserProfileResource


class AdminExportTest(TestCase):
    def setUp(self):
        self.client = Client()

        # Суперпользователь
        self.superuser = User.objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='admin123____'
        )

        # Пользователь + профиль
        self.user = User.objects.create_user(
            username='executor1',
            email='exec1@example.com',
            first_name='Алексей',
            last_name='Петров'
        )
        self.profile = UserProfile.objects.create(
            user=self.user,
            role='executor',
            specialization='Python Developer',
            portfolio='https://github.com/exec1'
        )

        # Заказ
        self.order = Order.objects.create(
            title='Разработка API',
            description='REST API на Django',
            customer=self.user,
            status='open',
            budget=15000.00
        )

        # Отклик
        self.bid = Bid.objects.create(
            order=self.order,
            executor=self.user,
            message='Готов выполнить!',
            price_proposal=12000.00
        )

        self.client.login(username='admin', password='admin123____')

    def test_userprofile_export_xlsx_structure(self):
        """Тест: экспорт UserProfile → правильные колонки и данные"""
        url = reverse('admin:accounts_userprofile_changelist')
        response = self.client.post(url, {
            '_selected_action': [str(self.profile.id)],
            'action': 'export_admin_action',
            'format': '1'  # ← ИСПРАВЛЕНО
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        expected_headers = [
            'id',
            'user__username',
            'user__email',
            'user__first_name',
            'user__last_name',
            'role',
            'specialization',
            'rating',
            'portfolio',
            'created_at',
            'updated_at'
        ]
        actual_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(expected_headers))]
        self.assertEqual(actual_headers, expected_headers)

        row2 = [ws.cell(row=2, column=i + 1).value for i in range(len(expected_headers))]
        self.assertEqual(row2[0], self.profile.id)
        self.assertEqual(row2[1], 'executor1')
        self.assertEqual(row2[2], 'exec1@example.com')
        self.assertEqual(row2[5], 'executor')
        self.assertEqual(row2[6], 'Python Developer')
        self.assertEqual(row2[8], 'https://github.com/exec1')

    def test_order_export_xlsx_contains_correct_data(self):
        """Тест: экспорт Order → содержит заказ и связанные данные"""
        url = reverse('admin:orders_order_changelist')
        response = self.client.post(url, {
            '_selected_action': [str(self.order.id)],
            'action': 'export_admin_action',  # ← ИСПРАВЛЕНО
            'format': '1'  # ← ИСПРАВЛЕНО
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        self.assertIn('customer__username', headers)
        self.assertIn('assigned_executor__username', headers)
        self.assertIn('budget', headers)

        row2 = [cell.value for cell in ws[2]]
        self.assertIn('executor1', row2)  # customer__username
        self.assertIn(15000.0, row2)      # budget

    def test_bid_export_xlsx_correct_columns(self):
        """Тест: экспорт Bid → правильные колонки"""
        url = reverse('admin:orders_bid_changelist')
        response = self.client.post(url, {
            '_selected_action': [str(self.bid.id)],
            'action': 'export_admin_action',  # ← ИСПРАВЛЕНО
            'format': '1'  # ← ИСПРАВЛЕНО
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        expected = [
            'id',
            'order__title',
            'executor__username',
            'executor__email',
            'message',
            'price_proposal',
            'created_at',
            'updated_at'
        ]
        actual = [cell.value for cell in ws[1]]
        self.assertEqual(actual, expected)

        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2[1], 'Разработка API')
        self.assertEqual(row2[2], 'executor1')
        self.assertEqual(row2[5], 12000.0)

    def test_simple_export(self):
        url = reverse('admin:accounts_userprofile_changelist')
        response = self.client.post(url, {
            '_selected_action': [str(self.profile.id)],
            'action': 'export_admin_action',
            'format': '1'
        })
        print("Response type:", response.get('Content-Type'))
        print("Response content (first 200 chars):", response.content[:200])
        self.assertEqual(response.status_code, 200)

    def test_userprofile_export_all_xlsx(self):
        url = reverse('admin:accounts_userprofile_changelist')
        response = self.client.post(url, {
            'format': '1',
            'resource': '',
            'userprofileresource_id': 'on',
            'userprofileresource_user__username': 'on',
            'userprofileresource_user__email': 'on',
            'userprofileresource_user__first_name': 'on',
            'userprofileresource_user__last_name': 'on',
            'userprofileresource_role': 'on',
            'userprofileresource_specialization': 'on',
            'userprofileresource_rating': 'on',
            'userprofileresource_portfolio': 'on',
            'userprofileresource_created_at': 'on',
            'userprofileresource_updated_at': 'on',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def test_userprofile_resource_export(self):
        dataset = UserProfileResource().export()
        self.assertEqual(dataset[0]['user__username'], 'executor1')